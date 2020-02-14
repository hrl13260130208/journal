import xlrd
import openpyxl
import re
import os
import logging
import time
import json
from journal.redis_manager import name_manager
from journal.common import Row_Name,common_article
import requests
from bs4 import BeautifulSoup
import uuid
import PyPDF2
from  xlutils import copy


logger = logging.getLogger("logger")

EXECEL_PATH=r"D:\execl\\"

sb={}
def create_colume_name_variable(path):
    '''
    创建表头变量（读取execl表头生成Row_Name中的变量）
    :param path:
    :return:
    '''
    rb = xlrd.open_workbook(path)
    r_sheet = rb.sheet_by_index(0)
    row0=r_sheet.row(0)
    for i in row0:
        print(i.value.upper()+"=\""+i.value+"\"")

def create_and_save_execel(section,*name):
    '''
    创建excel并将数据写入execl
    :param section:
    :return:
    '''

    if name.__len__() != 0:
        execel_name = create_execl_name(name[0])
    else:
        execel_name=create_execl_name(section)
    wb=openpyxl.Workbook()
    sheet=wb.create_sheet("sheet1",0)
    line=2
    write_first_line(sheet)
    while (True):
        article_data = name_manager().get_article_data(section)
        if article_data == None:
            break
        article = json.loads(article_data)
        sheet.cell(line,Row_Name.COLUME_NUM.get(Row_Name.EXCELNAME)+1,execel_name[:-5])
        sheet.cell(line,Row_Name.COLUME_NUM.get(Row_Name.SERIAL_NUMBER)+1,line-1)
        for key in article.keys():
            num=Row_Name.COLUME_NUM.get(key,None)
            if num != None:
                try:
                    sheet.cell(line,num+1,article[key])
                except:
                    print("错误:",line,num+1,article[key])

        line+=1

    wb.save(EXECEL_PATH+execel_name)


def create_execl_name(section):
    '''
    创建Excel的名称
    :param section:
    :return:
    '''
    date_time=time.strftime("%Y%m%d", time.localtime())
    return "wc_hrl_"+section+"_"+date_time+"_1_"+date_time+".xlsx"

def write_first_line(sheet):
    d=Row_Name.COLUME_NUM
    for key in d.keys():
        sheet.cell(1,d[key]+1,key)

def write_logs():
    '''
    创建日志
    :return:
    '''
    date_time = time.strftime("%Y%m%d", time.localtime())
    path=EXECEL_PATH+date_time
    if not os.path.exists(path):
        os.mkdir(path)
    journal_name="journal.txt"
    article_name="article.txt"
    write_log(1,path+"/"+journal_name)
    write_log(2,path+"/"+article_name)

def write_log(num,file_path):
    file=open(file_path,"a+",encoding="utf-8")
    nm=name_manager()
    while(True):

        if num ==1:
            temp_data=nm.get_journal_error_massage()
        elif num ==2:
            temp_data=nm.get_article_error_massage()
        if temp_data == None:
            break
        file.write(temp_data+"\n")

def update_by_num():
    execl = openpyxl.load_workbook("D:/test/zh.xlsx")
    sheet = execl.get_sheet_by_name("sheet1")
    for line in open("D:/test/num.txt").readlines():
        print(int(line))
        c=sheet.cell(row=int(line)+1, column = Row_Name.COLUME_NUM[Row_Name.ABSTRACT]+1)
        c.value=c.value+"."
    execl.save("D:/test/zh.xlsx")


def write_page_total(excel_path=None):
    com = common_article(None)
    execl = openpyxl.load_workbook(excel_path)
    sheet = execl.get_sheet_by_name("sheet1")

    for index, i in enumerate(sheet.rows):

        path = i[Row_Name.COLUME_NUM[Row_Name.BIO]+1].value
        page_total = i[Row_Name.COLUME_NUM[Row_Name.PAGE_TOTAL]].value
        print(index, page_total,path,os.path.exists(path))
        if page_total==None:
            if os.path.exists(path):
                try:
                    pages=com.checkpdf(path)
                    i[Row_Name.COLUME_NUM[Row_Name.PAGE_TOTAL]].value=pages
                except:
                    logger.error("执行出错！", exc_info=True)
    execl.save(excel_path)


def update_excel(excel_path=None):

    com=common_article(None)
    execl = openpyxl.load_workbook(excel_path)
    sheet = execl.get_sheet_by_name("sheet1")

    for index,i in enumerate(sheet.rows):
        print(index)

        excel_author = i[Row_Name.COLUME_NUM[Row_Name.AUTHOR_NAME]].value
        url = i[Row_Name.COLUME_NUM[Row_Name.ABS_URL]].value
        aff=i[Row_Name.COLUME_NUM[Row_Name.AFFILIATION]].value
        publisher=i[Row_Name.COLUME_NUM[Row_Name.PUBLISHER]].value


        if excel_author==Row_Name.AUTHOR_NAME:
            continue
        try:
            if publisher=="The Korean Society of Gastroenterology":
                if aff==None:
                    data = requests.get(url)
                    soup = BeautifulSoup(data.text, "html.parser")

                    aff_dict = {}
                    div = soup.find("div", {"id": "conBox"})
                    for p in div.find_all("p"):
                        sup = p.find("sup")
                        if sup != None:
                            key = sup.get_text().strip()
                            sup.extract()
                            aff_dict[key] = p.get_text().strip()
                        else:
                            aff_dict["0"] = p.get_text().strip()

                        p.extract()
                    print(aff_dict)
                    line = div.get_text()
                    line = line[:line.find("Correspondence")].strip()
                    au_dict = com.clear_authors(line, aff_dict.keys())
                    au, em, af = com.get_author_email_aff(au_dict, {}, aff_dict)

                    i[Row_Name.COLUME_NUM[Row_Name.AUTHOR_NAME]].value=au
                    i[Row_Name.COLUME_NUM[Row_Name.AFFILIATION]].value=af


        except:
            pass

    execl.save(excel_path)

def update_xls(excel_path=r"C:\public\目次采全文\1129doi\DOI.xls"):
    values=["ABS_URL","DOI"]
    header = {
        "User-Agent": "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.102 Safari/537.36"}
    rb = xlrd.open_workbook(excel_path)
    r_sheet = rb.sheet_by_index(0)
    wb = copy.copy(rb)
    w_sheet = wb.get_sheet(0)
    items={}
    list = r_sheet.row_values(0)
    for value in values:
        index = list.index(value)
        items[value]=index

    for row in range(r_sheet.nrows - 1):
        row_num=row+1
        url=r_sheet.cell(row_num,items[values[0]]).value
        doi=r_sheet.cell(row_num,items[values[1]]).value

        if doi=="":
            print(url,doi)
            try:

                soup = requests.get(url,headers=header,verify=False,timeout=60)

                bs_c=BeautifulSoup(soup.text,"html.parser")
                # print(bs_c)
                w_doi=None
                article_doi = bs_c.find("meta", {"scheme": "doi"})
                if article_doi != None:
                    w_doi = article_doi["content"]

                article_doi = bs_c.find("meta", {"name": "citation_doi"})
                if article_doi != None:
                    w_doi = article_doi["content"]
                article_doi = bs_c.find("meta", {"name": "DC.Identifier.DOI"})
                if article_doi != None:
                    w_doi = article_doi["content"]

                if w_doi!=None:
                    print(row_num,items[values[1]],w_doi)
                    w_sheet.write(row_num, items[values[1]], w_doi)
            except:
                pass

    wb.save(excel_path)

if __name__ == '__main__':
    # update_excel(excel_path="D:/test/wc_hrl_The Federation of American Societies for Experimental Biology_20191015_1_20191015.xlsx")
    update_xls()
    # write_page_total(excel_path=r"D:\test\wc_hrl_European Centre for Disease Prevention and Control_20191028_1_20191028.xlsx")

    # string="######d#####"
    # abs="Colorectal self-expanding metal stents have been widely used as a bridge to surgery in patients with acute malignant colonic obstruction by allowing a single- stage operation, or as a definitive palliative procedure in patients with inoperable tumors. Colonic stents are placed under either fluoroscopic or combined endoscopic and fluoroscopic guidance, with similar technical- success and complication rates. Placement of colonic stents is a very safe procedure with a low procedure-related mortality rate, but serious complications can develop and reinterventions are not uncommon. Most of the complications can be treated by minimally invasive or conservative techniques, while surgical interventions are required for most patients with perforation.Keywords: Colorectal cancer; Self-expandable metal stents; Complications"
    # num = abs.find("Keywords:")
    # print(num)
    # print(abs[:num])
    # download("http://www.iccm-central.org/Proceedings/ICCM13proceedings/SITE/PAPERS/Abstract-1536.pdf","C:/File/sdf.pdf")
    # data_2 = requests.get("https://www.arch-anim-breed.net/53/85/2010/")
    # bs_2 = BeautifulSoup(data_2.text, "html.parser")
    # print(bs_2)


